#counting words, symbols in text files and a total price of translation
import tkinter as tk
from tkinter import ttk
import tkinter.filedialog as fd
import tkinter.messagebox as msg
import tkinter.font as tkFont
import openpyxl as xl
from openpyxl.styles import Font
import docx

#clears all data
def clear_data():
    global symbols
    global words
    global symbolsNoSpaces
    global filesList
    global filesData
    symbols = 0
    words = 0
    symbolsNoSpaces = 0
    filesList = []
    filesData = {}
    statistics.delete(*statistics.get_children())
    symbolsInFile.set("")
    symbolsNoSpacesFile.set("")
    wordsInFile.set("")
    

#counts words, symbols in selected files    
def get_file():
    global symbols, words, symbolsNoSpaces
    global filesList, filesData
    clear_data()
    progress.set(0)
    filesList = fd.askopenfilenames(title="Choose text files", \
                    filetypes=[("All text files", ("*.docx", "*.txt")), ("Word files", ("*.docx")), ("Text-only files",("*.txt")),("All files","*.*")])
    progressbar["maximum"]=len(filesList)
    
    for file in filesList:
        if file.split(".")[-1] == "txt":
            currentWords=count_txt(file)['currentWords']
            currentSymbols = count_txt(file)['currentSymbols']
            currentNoSpaces= count_txt(file)['currentNoSpaces']
        elif file.split(".")[-1] == "docx":
            currentWords=count_docx(file)['currentWords']
            currentSymbols = count_docx(file)['currentSymbols']
            currentNoSpaces= count_docx(file)['currentNoSpaces']
            
        filesData[file]=[currentWords,currentSymbols,currentNoSpaces,0]
        words=words+currentWords
        symbols=symbols+currentSymbols
        symbolsNoSpaces=symbolsNoSpaces + currentNoSpaces
        progressbar.step()
        main.update()
        
    if price.get() != "":
        count_price()
    else:    
        for key in iter(filesData):
            statistics.insert("",'end',iid=key,text=key,values=filesData[key])
    refresh_totals()    
    
def count_txt(file):
    f=open(file, 'r')
    currentWords=0
    currentSymbols=0
    currentNoSpaces=0
    for line in f:
        wordslist = line.split()
        currentWords = currentWords + len(wordslist)
        currentSymbols = currentSymbols + len(line.rstrip("\n"))
        currentNoSpaces = currentNoSpaces + sum(len(item) for item in wordslist)
    f.close()
    return {'currentWords': currentWords, 
            'currentSymbols': currentSymbols, 
            'currentNoSpaces': currentNoSpaces}

def count_docx(file):
    doc = docx.Document(file)
    currentWords=0
    currentSymbols=0
    currentNoSpaces=0
    for para in doc.paragraphs:
        wordslist=para.text.split()
        currentWords = currentWords + len(wordslist)
        currentSymbols = currentSymbols + len(para.text)
        currentNoSpaces = currentNoSpaces + sum(len(item) for item in wordslist)
    
    return {'currentWords': currentWords, 
            'currentSymbols': currentSymbols, 
            'currentNoSpaces': currentNoSpaces}
    
def update_states(*args):
    global modeWord, modeSpace, modeNoSpace
    if v.get() == modeSpace:
        symbWithSpacesEntry.config(state='normal')
        symbWoutSpaces.set("")
        symbWoutSpacesEntry.config(state='disabled')
        update_button_count()
    elif v.get() == modeNoSpace:
        symbWithSpaces.set("")
        symbWithSpacesEntry.config(state='disabled')
        symbWoutSpacesEntry.config(state='normal')
        update_button_count()
    else:
        symbWithSpaces.set("")
        symbWithSpacesEntry.config(state='disabled')
        symbWoutSpaces.set("")
        symbWoutSpacesEntry.config(state='disable')
        update_button_count()
        
def refresh_statistics():
    global filesData
    statistics.delete(*statistics.get_children())
    for key in iter(filesData):
        statistics.insert("",1,text=key,values=filesData[key])

def refresh_totals():
    symbolsInFile.set(str(symbols))
    symbolsNoSpacesFile.set(str(symbolsNoSpaces))
    wordsInFile.set(str(words))
    finalPrice.set(str(totalPrice))
    
def count_price():
    global totalPrice
    global modeWord, modeSpace, modeNoSpace
    currentPrice=0
    try:
        clearPrice=float(price.get())
    except ValueError:
        msg.showerror("Error", "Wrong price value")
        priceEntry.focus()
        exit
    if v.get() == modeWord:
        clearSymbols = 1
    elif v.get() == modeSpace:
        try:
            clearSymbols=float(symbWithSpaces.get())
        except ValueError:
            msg.showerror("Error", "Wrong symbol value")
            symbWithSpacesEntry.focus()
            exit
    elif v.get() == modeNoSpaces:
        try:
            clearSymbols=float(symbWoutSpaces.get())
        except ValueError:
            msg.showerror("Error", "Wrong symbol value")
            symbWithSpacesEntry.focus()
            exit
    progress.set(0)
    progressbar['maximum']=len(filesData)
    for key in iter(filesData):
        filesData[key][3] = round(filesData[key][v.get()] / clearSymbols * clearPrice,2)
        progressbar.step()
        main.update()
    totalPrice = round(sum(filesData[key][3] for key in iter(filesData)),2)    
    refresh_totals()
    refresh_statistics()
    progress.set(0)
    main.update()
    excelButton.config(state='normal')
    

def update_button_count(*args):
    global modeWord, modeSpace, modeNoSpace
    excelButton.config(state='disable')
    if (price.get() != "") and filesData:        
        if v.get() == modeWord: 
            countButton.config(state='normal')
        elif (v.get() == modeSpace) and (symbWithSpaces.get() != ""):
            countButton.config(state='normal')            
        elif (v.get() == modeNoSpace) and (symbWoutSpaces.get() != ""):
            countButton.config(state='normal')                 
        else:
            countButton.config(state='disabled')                 
    else:
        countButton.config(state='disabled')

def export_to_excel():
    global columnTitles
    global filesData
    global modeWord, modeSpace, modeNoSpace
    wb = xl.Workbook()
    sh = wb.active
    sh.title = 'Totals'
    progressbar['maximum']=3
    progress.set(0)
    main.update()
    for i in range(len(columnTitles)):
        sh.cell(column=i+2, row=1, value=columnTitles[i])
        sh.cell(column=i+2, row=1).font=Font(bold=True)
    r=2
    progressbar.step()
    main.update()
    for key in iter(filesData):
        sh.cell(column=1, row=r, value=key)
        for i in range(len(filesData[key])):
            sh.cell(column=i+2, row=r, value=filesData[key][i])            
        r=r+1
    progressbar.step()
    main.update()
    comment = "Price is %s per " % price.get()
    if v.get() == modeWord:
        comment = comment + "1 word"
    elif v.get() == modeSpace:
        comment = comment + "%s symbols with spaces" \
                  % symbWithSpaces.get()
    elif v.get() == modeNoSpace:
        comment = comment + "%s symbols without spaces" \
                  %symbWoutSpaces.get()
    sh.cell(row=r+3, column=1, value=comment)
    progressbar.step()
    main.update()
    f = fd.asksaveasfile(mode='w', defaultextension=".xlsx", filetypes=(("Excel files", "*.xlsx"),("All Files", "*.*")))
    if f is None:
        return
    else:                
        wb.save(f.name)
        
#global variables
symbols = 0
words = 0
symbolsNoSpaces = 0
totalPrice = 0
columnTitles=['Words', 'Symbols with spaces',
              'Symbols w/out spaces', 'Price']
filesList=[]
filesData={}
modeWord = 0
modeSpace = 1
modeNoSpace = 2

#initialization of the main window
main=tk.Tk()
main.title("Translation price counter")
main.grid_columnconfigure(1, weight=1)
main.grid_rowconfigure(0, weight=1)
main.minsize(866,360)
#main.attributes('-zoomed', True)

mainframe=ttk.Frame(main, padding="10 10 10 10")
mainframe.grid(row=0,column=0,sticky=tk.W+tk.S+tk.E+tk.N)

fileframe=ttk.Frame(mainframe, padding="10 10 10 10")
fileframe.grid(row=0,column=0)

priceframe=ttk.Frame(mainframe,padding="0 10 0 0")
priceframe.grid(row=1,column=0, sticky=tk.W)
pricebuttonframe=ttk.Frame(priceframe)
pricebuttonframe.grid(row = 7, columnspan = 3)

resultframe=ttk.Frame(mainframe, padding="0 20 0 0")
resultframe.grid(row=2,column=0, sticky=tk.W+tk.S+tk.E+tk.N)

statframe=ttk.Frame(main, padding="10 10 10 10")
statframe.grid(row=0,column=1, sticky=tk.W+tk.S+tk.E+tk.N)
statframe.grid_rowconfigure(0,weight=1)
statframe.grid_columnconfigure(0,weight=1)

progressframe=ttk.Frame(main, padding="10 5 5 5", border=1, relief='sunken')
progressframe.grid(row=1, columnspan=2, sticky=tk.W + tk.S + tk.E + tk.N)

#===================== File section =============================
getFilePath=tk.StringVar()
getFilePath.set("")
getFileButton=ttk.Button(fileframe, text="Select text files", command=get_file)
getFileButton.grid(row=0,column=0)

#===================== Price section ============================

label2=ttk.Label(priceframe,text="Input price: ",)
label2.grid(row=0,columnspan=2,sticky=tk.W,padx="10 10")
price=tk.StringVar()
price.trace("w", update_button_count)
priceEntry=ttk.Entry(priceframe,width=6,textvariable=price)
priceEntry.grid(row=0,column=1,sticky=tk.W, padx="10 10")

v=tk.IntVar()
v.set(modeWord)
v.trace("w",update_states)
ttk.Radiobutton(priceframe, text="Per word", variable=v, value=modeWord).grid(row=2,column=0,sticky=tk.W,padx="10 10")

ttk.Radiobutton(priceframe, text="Per symbols with spaces", variable=v, value=modeSpace).grid(row=3,column=0,sticky=tk.W,padx="10 10")

symbWithSpaces=tk.StringVar()
symbWithSpaces.trace("w", update_states)
symbWithSpacesEntry=ttk.Entry(priceframe, width=6, textvariable=symbWithSpaces)
symbWithSpacesEntry.grid(column=1,row=3,sticky=tk.W,padx="10 10")
symbWithSpacesEntry.config(state='disabled')

ttk.Radiobutton(priceframe, text="Per symbols w/out spaces", variable=v, value=modeNoSpace).grid(row=4,column=0,sticky=tk.W,padx="10 10")
symbWoutSpaces=tk.StringVar()
symbWoutSpaces.trace("w", update_states)
symbWoutSpacesEntry=ttk.Entry(priceframe, width=6, textvariable=symbWoutSpaces)
symbWoutSpacesEntry.grid(column=1,row=4,sticky=tk.W,padx="10 10")
symbWoutSpacesEntry.config(state='disabled')

#count button
countButton=ttk.Button(pricebuttonframe,text="Count",command=count_price)
countButton.grid(row=0,column=0,pady=10, padx=10,sticky=tk.E)
countButton.config(state='disabled')
excelButton=ttk.Button(pricebuttonframe,text="Export to *.xls file", command = export_to_excel)
excelButton.grid(row=0,column=1, pady=10)
excelButton.config(state='disabled')
clearButton=ttk.Button(pricebuttonframe,text="Clear",command=clear_data)
clearButton.grid(row=0,column=2,pady=10,padx=10,sticky=tk.W)

#============================ Results section ====================
symbolsInFile=tk.StringVar()
symbolsInFile.set("")
wordsInFile=tk.StringVar()
wordsInFile.set("")

ttk.Label(resultframe,text=columnTitles[0]+": ").grid(column=0,row=0,sticky=tk.W)
wordsInFileLabel=ttk.Label(resultframe,text="",textvariable=wordsInFile).grid(column=1,row=0,sticky=tk.W)
ttk.Label(resultframe,text=columnTitles[1]+": ").grid(column=0,row=1,sticky=tk.W)
symbolsInFileLabel=ttk.Label(resultframe,text="",textvariable=symbolsInFile).grid(column=1,row=1,sticky=tk.W)
symbolsNoSpacesFile=tk.StringVar()
symbolsNoSpacesFile.set("")
ttk.Label(resultframe, text=columnTitles[2]+": ").grid(column=0,row=2,sticky=tk.W)
symbolsNoSpacesFileLabel=ttk.Label(resultframe,text="",textvariable=symbolsNoSpacesFile).grid(column=1, row=2,sticky=tk.W)

finalPrice=tk.StringVar()
ttk.Label(resultframe, text=columnTitles[3]+": ", font=tkFont.Font(weight='bold')) \
                       .grid(column=0,row=3,sticky=tk.W)
ttk.Label(resultframe, text="", textvariable=finalPrice, font=tkFont.Font(weight='bold')).grid(column=1,row=3,sticky=tk.W)

#========================== statistics section ===================
statistics=ttk.Treeview(statframe)
statistics.grid_rowconfigure(0,weight=1)
statistics.grid_columnconfigure(0,weight=1)
statistics.grid(row=0,column=0, sticky=tk.W+tk.S+tk.E+tk.N)
statistics["columns"]=("words","smbspaces","smbnospaces","price")
statistics.column("#0", minwidth=200, width=200, anchor=tk.W)
statistics.column("words", minwidth=50, width=50, anchor=tk.CENTER)
statistics.column("smbspaces", minwidth=100, width=100, anchor=tk.CENTER)
statistics.column("smbnospaces", minwidth=100, width=100, anchor=tk.CENTER)
statistics.column("price", minwidth=50, width=50, anchor=tk.CENTER)
statistics.heading("words",text=columnTitles[0])
statistics.heading("smbspaces", text=columnTitles[1])
statistics.heading("smbnospaces", text=columnTitles[2])
statistics.heading("price", text=columnTitles[3])

scrollbar=ttk.Scrollbar(statframe, orient='vertical', \
                        command = statistics.yview)
scrollbar.grid(row=0, column=1, sticky = tk.E + tk.N + tk.S)
statistics.configure(yscrollcommand=scrollbar.set)

#progress bar
progress=tk.IntVar()
ttk.Label(progressframe, text="Progress: ").grid(row=0, column=0)
progressbar=ttk.Progressbar(progressframe, variable=progress, orient="horizontal", \
                            mode="determinate", length=200)
progressbar.grid(row=0, column= 1)



main.mainloop()

